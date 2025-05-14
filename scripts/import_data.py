import pandas as pd
import sqlite3
import os
import re
import math
from datetime import datetime, date
import calendar
from argparse import ArgumentParser
from urllib.parse import unquote
from openpyxl import load_workbook
import traceback
import sys
from typing import Optional, Tuple, Dict, Any, List

# --- Column Mappings ---
COLUMN_MAPS = {
    "report": {
        "url": "URL",
        "full": "Full",
        "partial": "Partial",
        "total_bw": "Total BW",
        "avg_bw": "Avg BW"
    },
    "monthly": {
        "url": "Downloads",
        "full": "Hits",
        "partial": "206 Hits",
        "total_bw": "Bandwidth",
        "avg_bw": "Average size"
    }
}

# --- Expected DB Columns (after mapping, before adding calculated ones) ---
EXPECTED_MAPPED_COLS = {"url", "full", "partial", "total_bw", "avg_bw"}

def normalize_title_for_grouping_key(original_title, word_limit=2):
    if not original_title:
        return "_EMPTY_TITLE_"
    
    normalized = str(original_title).lower()
    
    # List of prefixes to strip (case-insensitive)
    prefixes_to_strip = ["id_", "iv_", "sp_"]
    for prefix in prefixes_to_strip:
        if normalized.startswith(prefix):
            normalized = normalized[len(prefix):]
            break # Assume only one such prefix will occur
    
    # Replace all non-alphanumeric characters with a single space, then strip leading/trailing spaces
    normalized = re.sub(r'[^a-z0-9]+', ' ', normalized).strip()
    
    # Split into words and take up to 'word_limit' words
    words = normalized.split() 
    significant_word_join = "".join(words[:word_limit]) 
    
    return significant_word_join if significant_word_join else "_PROCESSED_EMPTY_TITLE_"

def extract_code_feature_title(filename_url_part: str) -> Tuple[Optional[str], Optional[str], str, Optional[date]]:
    """
    Extract code, feature, title, and creation date from a filename or URL part.
    Returns a tuple of (code, feature, title, created_at).
    """
    if not filename_url_part:
        return None, None, "", None
    
    # Use the full input for date extraction
    processed_for_date = filename_url_part
    # Use only the filename for code/feature/title extraction
    if '/' in filename_url_part:
        processed_filename = filename_url_part.split('/')[-1]
    else:
        processed_filename = filename_url_part
    processed_filename = unquote(processed_filename).replace('.mp3', '').replace('.MP3', '')
    print(f"Processing filename: {processed_filename}")  # Debug print
    
    # Extract creation date from full input (YYYYMMDD format in filename or year/month in path)
    created_at = None
    if match := re.match(r'^(\d{8})', processed_filename):
        try:
            date_str = match.group(1)
            created_at = date(int(date_str[:4]), int(date_str[4:6]), int(date_str[6:8]))
            # Remove the date part from the filename for further processing
            processed_filename = processed_filename[8:]
        except ValueError:
            print(f"Warning: Could not parse date from filename prefix: {date_str}")
    # If no date in filename, try to extract from URL path
    if not created_at and '/' in processed_for_date:
        url_parts = processed_for_date.split('/')
        for part in url_parts:
            if re.match(r'^\d{4}$', part):  # Year
                year = int(part)
                # Look for month in next part
                month_idx = url_parts.index(part) + 1
                if month_idx < len(url_parts):
                    month_part = url_parts[month_idx]
                    if re.match(r'^\d{2}$', month_part):  # Month
                        try:
                            month = int(month_part)
                            created_at = date(year, month, 1)  # Use first day of month
                            break
                        except ValueError:
                            continue
    # Pattern 1: DDD@FEATURE_OptionalSeparator_TITLE
    if match := re.match(r'^(\d{3})@(HPCpodcast|HPCNB|Mktg_Podcast)', processed_filename):
        code_val = match.group(1)
        feature_val = match.group(2)
        title_val = processed_filename[match.end():]
        if title_val.startswith('_'):
            title_val = title_val[1:]
        print(f"Pattern 1 match - Code: {code_val}, Feature: {feature_val}, Title: {title_val}")
        return code_val, feature_val, title_val, created_at
    # Pattern 2: ADDD-TITLE
    if match := re.match(r'^(A\d{3})-', processed_filename):
        code_val = match.group(1)
        title_val = processed_filename[match.end():]
        print(f"Pattern 2 match - Code: {code_val}, Title: {title_val}")
        return code_val, None, title_val, created_at
    # Pattern 3: FEATURE_CODE_TITLE (e.g. HPC001_TITLE or OXD001_TITLE)
    if match := re.match(r'^([A-Z]+)(\d{3})_', processed_filename):
        feature_part = match.group(1)
        code_part = match.group(2)
        title_part = processed_filename[match.end():]
        print(f"Pattern 3 match - Code: {code_part}, Feature: {feature_part}, Title: {title_part}")
        return code_part, feature_part, title_part, created_at
    # Pattern 4: FEATURE_TITLE or FEATURE_CODE_TITLE
    if match_main_feature := re.match(r'^(HPCpodcast|HPCNB|Mktg_Podcast|OXD)_', processed_filename):
        feature_val = match_main_feature.group(1)
        temp_title = processed_filename[match_main_feature.end():]
        if match_code_and_title := re.match(r'^(\d{3})_(.*)', temp_title):
            code_val = match_code_and_title.group(1)
            final_title = match_code_and_title.group(2)
            print(f"Pattern 4a match - Code: {code_val}, Feature: {feature_val}, Title: {final_title}")
            return code_val, feature_val, final_title, created_at
        elif match_code_only := re.match(r'^(\d{3})$', temp_title):
            code_val = match_code_only.group(1)
            print(f"Pattern 4b match - Code: {code_val}, Feature: {feature_val}, Title: ''")
            return code_val, feature_val, "", created_at
        else:
            print(f"Pattern 4c match - Feature: {feature_val}, Title: {temp_title}")
            return None, feature_val, temp_title, created_at
    # Fallback: no specific pattern matched
    print(f"No pattern match - Using full filename as title: {processed_filename}")
    return None, None, processed_filename, created_at

def extract_created_at_from_url(url):
    match = re.search(r"/(\d{4})/(\d{2})/", url)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), 1) # Day is 1
        except ValueError:
            return None
    return None

def parse_excel_filename_date(filename_only):
    match = re.match(r"(\d{4})(\d{2})(\d{2})_.*\.xlsx?", filename_only, re.IGNORECASE)
    if match:
        year, month, day_val = int(match.group(1)), int(match.group(2)), int(match.group(3))
        # day_val from filename is ignored for setting the date, we use last day of month
        try:
            _, last_day_of_month = calendar.monthrange(year, month)
            return date(year, month, last_day_of_month), year, month
        except ValueError:
            print(f"Warning: Could not parse date from filename {filename_only} with parts {year}-{month}-(last_day)")
            return None, None, None
    return None, None, None

def format_bw(value_mb):
    if value_mb is None or pd.isna(value_mb):
        return "N/A"
    if value_mb >= 1024:
        return f"{value_mb/1024:.2f} GB"
    return f"{value_mb:.2f} MB"

def print_aggregation_summary_for_sheet(sheet_name, aggregated_data, consumed_year, consumed_month_display, file_type):
    aggregation_found = False
    for data_item in aggregated_data.values(): # Iterate directly over values
        if data_item["num_records_aggregated"] > 1:
            aggregation_found = True
            break
    
    if not aggregation_found:
        return

    print(f"\n--- Intra-File Duplicate Aggregation Summary Table for Sheet: '{sheet_name}' ({file_type} data) ---")
    # Define headers for the new single table
    headers = [
        "No.", "URL (Last 45 Chars)", "Period", "Orig.", "Title (First)",
        "Sum Full", "Sum Part", "Sum Total BW", "Calc Eq.Full", "Calc Avg.BW"
    ]
    # Define column widths (approximate, adjust as needed)
    col_widths = [3, 45, 10, 5, 60, 8, 8, 14, 12, 12] # Increased Title column width to 60

    header_line = " | ".join([h.ljust(w) for h, w in zip(headers, col_widths)])
    print(header_line)
    print("-" * len(header_line))

    count = 0
    for key, data_item in aggregated_data.items():
        if data_item["num_records_aggregated"] > 1:
            count += 1
            url_display = data_item['url']
            if len(url_display) > col_widths[1]: # URL column width
                url_display = "..." + url_display[-(col_widths[1]-3):]
            
            consumed_period_str = f"{consumed_year}-{consumed_month_display:02d}" if consumed_month_display else str(consumed_year)
            num_original_rows = str(data_item['num_records_aggregated'])
            title_display = data_item['title']
            
            eq_full_display = (math.floor(data_item["full_sum"] + 0.5 * data_item["partial_sum"]) 
                             if pd.notnull(data_item["full_sum"]) and pd.notnull(data_item["partial_sum"]) else None)
            
            avg_bw_display = None
            if (data_item["full_sum"] + data_item["partial_sum"]) > 0:
                avg_bw_display = data_item["total_bw_sum"] / (data_item["full_sum"] + data_item["partial_sum"])

            row_values = [
                str(count),
                url_display,
                consumed_period_str,
                num_original_rows,
                title_display,
                str(int(data_item['full_sum'])),
                str(int(data_item['partial_sum'])),
                format_bw(data_item['total_bw_sum']),
                str(int(eq_full_display)) if eq_full_display is not None else 'N/A',
                format_bw(avg_bw_display)
            ]
            
            # Ensure all row_values are strings before ljust
            row_values_str = [str(v) for v in row_values]
            row_line = " | ".join([val.ljust(w) for val, w in zip(row_values_str, col_widths)])
            print(row_line)
            
    if aggregation_found:
      print("--- End of Aggregation Summary Table ---")

def parse_float(val):
    if isinstance(val, str):
        val = val.upper().replace("MB", "").strip()
        if "GB" in val:
            val = val.replace("GB", "").strip()
            try:
                return float(val) * 1024
            except (ValueError, TypeError):
                return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

MONTHLY_FILENAME_PATTERN = re.compile(r"^\d{8}_podcast_downloads\.xlsx?$", re.IGNORECASE)
PREFIXED_TIMESTAMP_PATTERN = re.compile(r"^\d{8}-\d{6}_(.*)", re.IGNORECASE)
APPENDED_TIMESTAMP_PATTERN = re.compile(r"^(.*?)_(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})(\.(?:xlsx?|xls))$", re.IGNORECASE)

def detect_file_type(filename):
    # print(f"\n[DEBUG detect_file_type] Received filename: '{filename}'")

    # 1. Try to detect based on the new appended timestamp format
    match_appended_ts = APPENDED_TIMESTAMP_PATTERN.match(filename)
    if match_appended_ts:
        original_base = match_appended_ts.group(1)
        # timestamp_part = match_appended_ts.group(2) # Not needed for logic, only for debug
        original_ext = match_appended_ts.group(3)
        filename_to_test = f"{original_base}{original_ext}"
        # print(f"  [DEBUG detect_file_type] Appended TS matched. Original Base: '{original_base}', Timestamp: '{timestamp_part}', Ext: '{original_ext}'")
        # print(f"  [DEBUG detect_file_type] Testing with (appended logic): '{filename_to_test}'")

        if MONTHLY_FILENAME_PATTERN.match(filename_to_test):
            # print(f"    [DEBUG detect_file_type] Matched MONTHLY_FILENAME_PATTERN on '{filename_to_test}'. Type: monthly")
            return "monthly"
        if filename_to_test.lower().startswith("report") and filename_to_test.lower().endswith(('.xls', '.xlsx')):
            # print(f"    [DEBUG detect_file_type] Matched report pattern on '{filename_to_test}'. Type: report")
            return "report"
        # print(f"    [DEBUG detect_file_type] '{filename_to_test}' did not match monthly or report after stripping appended TS.")
    # else:
        # print(f"  [DEBUG detect_file_type] Appended TS pattern did NOT match.")

    # 2. If not matched above OR if type not determined, try prefixed timestamp format
    match_prefixed_ts = PREFIXED_TIMESTAMP_PATTERN.match(filename)
    if match_prefixed_ts:
        original_filename_part_after_prefix = match_prefixed_ts.group(1)
        # print(f"  [DEBUG detect_file_type] Prefixed TS matched. Testing with: '{original_filename_part_after_prefix}'")
        if MONTHLY_FILENAME_PATTERN.match(original_filename_part_after_prefix):
            # print(f"    [DEBUG detect_file_type] Matched MONTHLY_FILENAME_PATTERN on '{original_filename_part_after_prefix}'. Type: monthly")
            return "monthly"
        if original_filename_part_after_prefix.lower().startswith("report") and original_filename_part_after_prefix.lower().endswith(('.xls', '.xlsx')):
            # print(f"    [DEBUG detect_file_type] Matched report pattern on '{original_filename_part_after_prefix}'. Type: report")
            return "report"
        # print(f"    [DEBUG detect_file_type] '{original_filename_part_after_prefix}' did not match monthly or report after stripping prefixed TS.")
    # else:
        # print(f"  [DEBUG detect_file_type] Prefixed TS pattern did NOT match.")

    # 3. If no known timestamp format matched OR if type not determined, try on the filename as-is
    # print(f"  [DEBUG detect_file_type] Trying to match on original filename as-is: '{filename}'")
    if MONTHLY_FILENAME_PATTERN.match(filename):
        # print(f"    [DEBUG detect_file_type] Matched MONTHLY_FILENAME_PATTERN on original '{filename}'. Type: monthly")
        return "monthly"
    if filename.lower().startswith("report") and filename.lower().endswith(('.xls', '.xlsx')):
        # print(f"    [DEBUG detect_file_type] Matched report pattern on original '{filename}'. Type: report")
        return "report"
    
    # print(f"[DEBUG detect_file_type] All checks failed. Returning None for '{filename}'.")
    return None

def get_actual_hyperlink_url(worksheet, cell):
    """Extract the actual hyperlink URL from an Excel cell."""
    if cell.hyperlink:
        return cell.hyperlink.target
    return cell.value

def read_excel_with_hyperlinks(filepath: str, sheet_name: str) -> pd.DataFrame:
    """
    Reads an Excel sheet, extracting hyperlink URLs or cell values for 'URL' or 'Downloads' columns, and filters for relevant podcast rows.
    Accepts both full URLs and relative paths containing '/wp-content/uploads'.
    """
    try:
        print(f"\nProcessing sheet: {sheet_name}")
        wb = load_workbook(filepath, data_only=False)  # data_only=False to preserve hyperlinks
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in {filepath}")
            return pd.DataFrame()
        ws = wb[sheet_name]
    except FileNotFoundError:
        print(f"Error: File not found at {filepath}")
        return pd.DataFrame()
    except Exception as e:
        print(f"Error loading workbook {filepath}: {e}")
        traceback.print_exc()
        return pd.DataFrame()

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        header_value = cell.value
        if header_value is None:
            header_value = ""
        headers.append(str(header_value).strip())
    
    print(f"Found headers: {headers}")
    data = []
    
    # Find URL column index
    url_col_idx = None
    for idx, header in enumerate(headers):
        if header in ["URL", "Downloads"]:
            url_col_idx = idx
            print(f"Found URL column at index {idx}: {header}")
            break
    
    if url_col_idx is None:
        print(f"Warning: No URL/Downloads column found in sheet '{sheet_name}'")
        return pd.DataFrame()

    # Process rows
    row_count = 0
    skipped_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            url_cell = row[url_col_idx]
            url = None
            
            # Try to get URL from hyperlink first
            if url_cell.hyperlink:
                url = url_cell.hyperlink.target
                print(f"Row {row_idx}: Found hyperlink: {url}")
            # Fallback to cell value (relative or full path)
            elif url_cell.value and isinstance(url_cell.value, str):
                potential_url = str(url_cell.value).strip()
                if "/wp-content/uploads" in potential_url:
                    url = potential_url
                    print(f"Row {row_idx}: Using cell value as URL: {url}")
            
            if not url or "/wp-content/uploads" not in url:
                skipped_count += 1
                continue
                
            row_data = {}
            for header, cell in zip(headers, row):
                if header in ["URL", "Downloads"]:
                    row_data[header] = url
                else:
                    # Convert cell value to appropriate type
                    if cell.value is None:
                        row_data[header] = None
                    elif isinstance(cell.value, (int, float)):
                        row_data[header] = cell.value
                    else:
                        row_data[header] = str(cell.value).strip()
            
            data.append(row_data)
            row_count += 1
            
        except Exception as e:
            print(f"Error processing row {row_idx} in sheet '{sheet_name}': {e}")
            traceback.print_exc()
            continue

    if not data:
        print(f"Info: No valid podcast URLs found in sheet '{sheet_name}'")
        print(f"Total rows skipped: {skipped_count}")
        return pd.DataFrame()

    df = pd.DataFrame(data)
    df.columns = df.columns.str.strip()
    
    # Ensure URL column is string type and not truncated
    if "URL" in df.columns:
        df["URL"] = df["URL"].astype(str)
        print(f"URL column sample (first 5 rows):\n{df['URL'].head().to_string()}")
    elif "Downloads" in df.columns:
        df["Downloads"] = df["Downloads"].astype(str)
        print(f"Downloads column sample (first 5 rows):\n{df['Downloads'].head().to_string()}")
    
    print(f"Successfully processed {row_count} rows from sheet '{sheet_name}'")
    print(f"Total rows skipped: {skipped_count}")
    return df

def import_data(filepath: str, override: bool = False, dry_run: bool = False, reset_db: bool = False) -> dict:
    """
    Import podcast data from Excel files into SQLite database.
    
    Args:
        filepath: Path to the Excel file
        override: Whether to override existing database
        dry_run: Whether to perform a dry run (no actual database changes)
        reset_db: Whether to reset the database before import
    """
    db_path = "data/podcasts.db"
    filename_only = os.path.basename(filepath)
    file_type = detect_file_type(filename_only)

    if not file_type:
        print(f"Error: Could not automatically determine file type for '{filename_only}'.")
        return

    if reset_db and os.path.exists(db_path) and not dry_run:
        print(f"🗑️ Removing existing database {db_path} due to --reset-db flag.")
        os.remove(db_path)

    # Create database and table
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("""
    CREATE TABLE IF NOT EXISTS podcasts (
        url TEXT NOT NULL,
        title TEXT,
        code TEXT,
        feature TEXT,
        full INTEGER,
        partial INTEGER,
        avg_bw REAL,
        total_bw REAL,
        eq_full INTEGER,
        created_at TEXT,        
        consumed_at TEXT,       
        consumed_year INTEGER NOT NULL, 
        consumed_month INTEGER NOT NULL,
        assumed_month INTEGER NOT NULL DEFAULT 0,
        imported_at TEXT,
        source_file_path TEXT,
        PRIMARY KEY (url, consumed_year, consumed_month)
    )
    """)
    conn.commit()

    try:
        wb = load_workbook(filepath, read_only=True, data_only=False)
        all_sheet_names = wb.sheetnames
        initial_sheets_in_file = len(all_sheet_names)
        print(f"\nFound {initial_sheets_in_file} sheets in file: {all_sheet_names}")
    except Exception as e:
        print(f"Error reading file {filepath}: {e}")
        traceback.print_exc()
        conn.close()
        return

    # Initialize statistics
    stats = {
        'dry_run': {'inserted': 0, 'replaced': 0, 'ignored': 0},
        'actual': {'inserted': 0, 'replaced': 0, 'ignored': 0},
        'sheets': {
            'total': len(all_sheet_names),
            'processed': 0,
            'skipped': {'unreadable': 0, 'missing_cols': 0, 'bad_date': 0}
        },
        'rows': {
            'scanned': 0,
            'merged': 0,
            'errors': 0
        }
    }

    # Determine which sheets to process
    sheets_to_process = all_sheet_names if file_type == "report" else [all_sheet_names[0]] if all_sheet_names else []
    stats['sheets']['total'] = len(sheets_to_process)
    print(f"\nProcessing {len(sheets_to_process)} sheets: {sheets_to_process}")

    for sheet_name in sheets_to_process:
        try:
            print(f"\n{'='*50}")
            print(f"Processing sheet: {sheet_name}")
            print(f"{'='*50}")
            
            df = read_excel_with_hyperlinks(filepath, sheet_name)
            if df.empty:
                print(f"No data found in sheet '{sheet_name}'")
                continue

            # Map columns according to file type
            col_map = COLUMN_MAPS[file_type]
            df.columns = [str(col).strip() for col in df.columns]
            renamed_cols = {v: k for k, v in col_map.items() if v in df.columns}
            df.rename(columns=renamed_cols, inplace=True)

            if not EXPECTED_MAPPED_COLS.issubset(df.columns):
                print(f"Warning: Sheet '{sheet_name}' missing required columns. Expected: {EXPECTED_MAPPED_COLS}. Got: {set(df.columns)}")
                stats['sheets']['skipped']['missing_cols'] += 1
                continue

            # Determine consumption date
            if file_type == "report":
                try:
                    consumed_year = int(sheet_name)
                    consumed_month = 12
                    consumed_at = date(consumed_year, 12, 31).isoformat()
                    assumed_month = 1
                    print(f"Using year {consumed_year} from sheet name")
                except ValueError:
                    print(f"Warning: Invalid year in sheet name '{sheet_name}'")
                    stats['sheets']['skipped']['bad_date'] += 1
                    continue
            else:  # monthly
                parsed_date, yr, mn = parse_excel_filename_date(filename_only)
                if not parsed_date:
                    print(f"Warning: Could not parse date from filename {filename_only}")
                    stats['sheets']['skipped']['bad_date'] += 1
                    continue
                consumed_year, consumed_month = yr, mn
                consumed_at = parsed_date.isoformat()
                assumed_month = 0
                print(f"Using date from filename: {consumed_at}")

            # Process rows
            stats['rows']['scanned'] += len(df)
            aggregated_data = {}
            reconciliation_log = []  # Track reconciliation events for reporting

            # Debug: Print the full extracted title for the first 10 rows
            debug_title_count = 0
            for _, row in df.iterrows():
                url = row['url']
                code, feature, title, created_at = extract_code_feature_title(url)
                if debug_title_count < 10:
                    print(f"[DEBUG] Extracted title: '{title}' from filename: '{url}")
                    debug_title_count += 1

            for _, row in df.iterrows():
                try:
                    url = row['url']
                    if pd.isna(url) or not isinstance(url, str) or not url.strip():
                        continue

                    # Extract metadata from URL
                    code, feature, title, created_at = extract_code_feature_title(url)
                    created_at_str = created_at.isoformat() if created_at else None

                    # Calculate metrics
                    full = pd.to_numeric(row.get('full'), errors='coerce')
                    partial = pd.to_numeric(row.get('partial'), errors='coerce')
                    total_bw = parse_float(row.get('total_bw'))
                    avg_bw = parse_float(row.get('avg_bw'))

                    # Create aggregation key (canonicalized)
                    agg_key = (
                        str(code) if code else "_NO_CODE_",
                        str(feature) if feature else "_NO_FEATURE_",
                        normalize_title_for_grouping_key(title),
                        consumed_year,
                        consumed_month
                    )

                    # Track all variants for reconciliation
                    if agg_key not in aggregated_data:
                        aggregated_data[agg_key] = {
                            'urls': [url],
                            'titles': [title],
                            'code': code,
                            'feature': feature,
                            'created_at': created_at_str,
                            'full_sum': float(full) if pd.notna(full) else 0.0,
                            'partial_sum': float(partial) if pd.notna(partial) else 0.0,
                            'total_bw_sum': float(total_bw) if pd.notna(total_bw) else 0.0,
                            'count': 1
                        }
                    else:
                        agg = aggregated_data[agg_key]
                        agg['urls'].append(url)
                        agg['titles'].append(title)
                        agg['full_sum'] += float(full) if pd.notna(full) else 0.0
                        agg['partial_sum'] += float(partial) if pd.notna(partial) else 0.0
                        agg['total_bw_sum'] += float(total_bw) if pd.notna(total_bw) else 0.0
                        agg['count'] += 1

                except Exception as e:
                    print(f"Error processing row in sheet '{sheet_name}': {e}")
                    traceback.print_exc()
                    stats['rows']['errors'] += 1
                    continue

            # Insert aggregated data into database
            imported_at = datetime.now().isoformat()
            for agg_key, agg_data in aggregated_data.items():
                try:
                    if agg_data['count'] > 1:
                        stats['rows']['merged'] += agg_data['count'] - 1

                    # Choose the most complete (longest) title/url for canonicalization
                    canonical_title = max(agg_data['titles'], key=len)
                    canonical_url = max(agg_data['urls'], key=len)

                    # --- Ensure title is always stripped of audio file extensions for consistency ---
                    canonical_title_clean = re.sub(r'\.(mp3|wav|aac|m4a)$', '', canonical_title, flags=re.IGNORECASE)

                    # Reconciliation reporting: If more than one variant, log the merge
                    if len(set(agg_data['titles'])) > 1 or len(set(agg_data['urls'])) > 1:
                        reconciliation_log.append({
                            'agg_key': agg_key,
                            'titles': list(set(agg_data['titles'])),
                            'urls': list(set(agg_data['urls'])),
                            'canonical_title': canonical_title_clean,
                            'canonical_url': canonical_url
                        })

                    # Calculate derived metrics
                    eq_full = math.floor(agg_data['full_sum'] + 0.5 * agg_data['partial_sum'])
                    avg_bw = (agg_data['total_bw_sum'] / (agg_data['full_sum'] + agg_data['partial_sum'])
                             if (agg_data['full_sum'] + agg_data['partial_sum']) > 0 else None)

                    db_values = (
                        canonical_url,
                        canonical_title_clean,
                        agg_data['code'],
                        agg_data['feature'],
                        agg_data['full_sum'],
                        agg_data['partial_sum'],
                        avg_bw,
                        agg_data['total_bw_sum'],
                        eq_full,
                        agg_data['created_at'],
                        consumed_at,
                        consumed_year,
                        consumed_month,
                        assumed_month,
                        imported_at,
                        filepath
                    )

                    if dry_run:
                        c.execute("SELECT 1 FROM podcasts WHERE url = ? AND consumed_year = ? AND consumed_month = ?",
                                  (canonical_url, consumed_year, consumed_month))
                        exists = c.fetchone()
                        if file_type == "monthly":
                            stats['dry_run']['replaced' if exists else 'inserted'] += 1
                        else:
                            stats['dry_run']['ignored' if exists else 'inserted'] += 1
                    else:
                        if file_type == "monthly":
                            c.execute("SELECT 1 FROM podcasts WHERE url = ? AND consumed_year = ? AND consumed_month = ?",
                                    (canonical_url, consumed_year, consumed_month))
                            exists = c.fetchone()
                            c.execute("""
                                INSERT OR REPLACE INTO podcasts (
                                    url, title, code, feature, full, partial, avg_bw, total_bw, eq_full,
                                    created_at, consumed_at, consumed_year, consumed_month, assumed_month, imported_at, source_file_path
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, db_values)
                            stats['actual']['replaced' if exists else 'inserted'] += 1
                        else:
                            c.execute("""
                                INSERT OR IGNORE INTO podcasts (
                                    url, title, code, feature, full, partial, avg_bw, total_bw, eq_full,
                                    created_at, consumed_at, consumed_year, consumed_month, assumed_month, imported_at, source_file_path
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, db_values)
                            if c.rowcount > 0:
                                stats['actual']['inserted'] += 1
                            else:
                                stats['actual']['ignored'] += 1

                except Exception as e:
                    print(f"Error inserting aggregated data for URL {agg_data['urls']}: {e}")
                    traceback.print_exc()
                    stats['rows']['errors'] += 1
                    continue

            stats['sheets']['processed'] += 1
            print(f"Successfully processed sheet '{sheet_name}'")

            # Print reconciliation summary for this sheet if any merges occurred
            if reconciliation_log:
                print("\n[RECONCILIATION SUMMARY]")
                for rec in reconciliation_log:
                    print(f"Merged variants for key {rec['agg_key']}:")
                    print(f"  Titles: {rec['titles']}")
                    print(f"  URLs: {rec['urls']}")
                    print(f"  Canonical Title: {rec['canonical_title']}")
                    print(f"  Canonical URL: {rec['canonical_url']}")
                print("[END RECONCILIATION SUMMARY]\n")

        except Exception as e:
            print(f"Error processing sheet '{sheet_name}': {e}")
            traceback.print_exc()
            stats['sheets']['skipped']['unreadable'] += 1
            continue

    conn.commit()
    conn.close()

    # Print summary
    print("\n" + "="*70)
    print(f" Import Summary for: {filename_only}")
    print("="*70)
    print(f"File Type: {file_type.capitalize()}")
    print(f"Processing Mode: {'Dry Run' if dry_run else 'Actual Import'}")
    print(f"Source Path: {filepath}\n")
    
    print("File & Sheet Details:")
    print(f"  Sheets in file: {initial_sheets_in_file}")
    print(f"  Sheets targeted: {stats['sheets']['total']}")
    print(f"  Sheets processed: {stats['sheets']['processed']}")
    print("  Sheets skipped:")
    print(f"    - Unreadable: {stats['sheets']['skipped']['unreadable']}")
    print(f"    - Missing columns: {stats['sheets']['skipped']['missing_cols']}")
    print(f"    - Bad date format: {stats['sheets']['skipped']['bad_date']}\n")
    
    print("Row Processing:")
    print(f"  Total rows scanned: {stats['rows']['scanned']}")
    print(f"  Rows merged: {stats['rows']['merged']}")
    print(f"  Processing errors: {stats['rows']['errors']}\n")
    
    if dry_run:
        print("Database Changes (Preview):")
        print(f"  Would insert: {stats['dry_run']['inserted']}")
        print(f"  Would replace: {stats['dry_run']['replaced']}")
        print(f"  Would ignore: {stats['dry_run']['ignored']}")
    else:
        print("Database Changes (Actual):")
        print(f"  Inserted: {stats['actual']['inserted']}")
        print(f"  Replaced: {stats['actual']['replaced']}")
        print(f"  Ignored: {stats['actual']['ignored']}")
    
    print("-"*70)

    return stats

if __name__ == "__main__":
    parser = ArgumentParser(description="Import podcast data from Excel files into SQLite database.") # Keep description generic or update
    parser.add_argument("filepath", help="Path to Excel file") # Or more generic "Path to data file"
    parser.add_argument("--override-db", action="store_true", help="Delete and recreate the database before import. Use with caution.")
    parser.add_argument("--dry-run", action="store_true", help="Preview actions only; no changes will be made to the database.")
    parser.add_argument("--reset-db", action="store_true", help="Reset the database before import. Use with caution.")
    args = parser.parse_args()
    import_data(args.filepath, args.override_db, args.dry_run, args.reset_db) # Call renamed function 